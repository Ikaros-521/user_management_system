from fastapi import APIRouter, HTTPException, status, Depends, Request, UploadFile, File, Form
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt, ExpiredSignatureError
from sqlalchemy import select, update, delete, and_, func
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime, timedelta
from loguru import logger
import traceback
import json

from openpyxl import load_workbook
from io import BytesIO

from utils.database import database, users_table, transaction_manager
from utils.models import UserResult, UserResultData, RefreshTokenRequest, TokenResponse, TokenData
from utils.models import UpdatePasswordMessage, UpdateUserInfoMessage, GetUserMessage, CommonResult
from utils.models import UserListData, UserListResult, AddUserMessage, UpdateUserMessage, DelUserMessage, DelUsersMessage
from utils.common import format_datetime, dict_to_string, role_name_converter
from utils.db_common import get_user_roles, insert_msg_to_logs_table, check_super_admin_role
from utils.config import local_config
from utils.security import verify_password, create_access_token, hash_password, validate_username, validate_email, validate_password, validate_phone

router = APIRouter()

SECRET_KEY = "mysecretkey"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_MINUTES = 60

user_sessions = {}
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")



# 校验token是否有效
async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            logger.error("Token payload does not contain 'sub'")
            raise credentials_exception
    except JWTError as e:
        logger.error(f"JWT decode error: {e}")
        raise credentials_exception
    user = await database.fetch_one(select(users_table).where(users_table.c.username == username))
    if user is None:
        logger.error(f"User not found for username: {username}")
        raise credentials_exception
    # logger.debug(f"User authenticated: {username}")
    return user

# 判断用户是否被禁用、token是否有效
async def get_current_active_user(current_user: dict = Depends(get_current_user)):
    if current_user["disabled"] == 1:
        logger.error(f"账号：{current_user['username']} 被禁用")
        raise HTTPException(status_code=400, detail="账号被禁用")
    # logger.debug(f"可用账号: {current_user['username']}")
    return current_user

@router.post("/auth/login", response_model=UserResult)
async def login_for_access_token(request: Request):
    try:
        form_data = await request.json()
        username = form_data.get("username")
        password = form_data.get("password")
        
        log_msg = ""
        
        if not validate_username(username)["ret"] or not validate_password(password)["ret"]:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="用户名或密码格式错误，请检查",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        query = select(users_table).where(users_table.c.username == username)
        user = await database.fetch_one(query)

        if not user or not verify_password(password, user["hashed_password"]):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Incorrect username or password",
                headers={"WWW-Authenticate": "Bearer"},
            )
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        refresh_token_expires = timedelta(minutes=REFRESH_TOKEN_EXPIRE_MINUTES)
        
        access_token = create_access_token(
            data={"sub": user["username"]}, expires_delta=access_token_expires
        )
        refresh_token = create_access_token(
            data={"sub": user["username"], "type": "refresh"}, expires_delta=refresh_token_expires
        )

        # 检查并处理重复登录
        if username in user_sessions:
            old_token = user_sessions[username]
            log_msg = f"用户 {username} 已经登录，踢掉旧会话"
            logger.info(log_msg)
            await insert_msg_to_logs_table({"username": username}, log_msg)
        
        user_sessions[username] = access_token

        log_msg = f"用户 {username} 登录成功"
        await insert_msg_to_logs_table({"username": username}, log_msg)
        
        # 判断是否已经过期
        if datetime.now() > user["expiration_ts"]:
            return UserResult(code=0, success=False, data=UserResultData(
                avatar=user["avatar"],
                username=user["username"],
                nickname=user["nickname"],
                roles=user["roles"].split(","),
                accessToken=access_token,
                refreshToken=refresh_token,
                expiration_ts=user["expiration_ts"],
                expires=datetime.utcnow() + access_token_expires
            ))

        return UserResult(code=0, success=True, data=UserResultData(
            avatar=user["avatar"],
            username=user["username"],
            nickname=user["nickname"],
            roles=user["roles"].split(","),
            accessToken=access_token,
            refreshToken=refresh_token,
            expiration_ts=user["expiration_ts"],
            expires=datetime.utcnow() + access_token_expires
        ))
    except Exception as e:
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=str(e),
            headers={"WWW-Authenticate": "Bearer"},
        )

@router.post("/auth/check_expiration")
async def check_expiration(current_user: dict = Depends(get_current_active_user)):
    try:
        # 构建基本查询
        query = select(users_table).where(users_table.c.username == current_user["username"])
        user = await database.fetch_one(query)

        if user is None:
            return CommonResult(code=0, success=False, data={"msg": "账号不存在"})

        # 判断是否已经过期
        if datetime.now() > user["expiration_ts"]:
            return CommonResult(code=0, success=False, data={"msg": "账号已过期", "expiration_ts": user["expiration_ts"]})
        else:
            return CommonResult(code=0, success=True, data={"msg": "账号可用", "expiration_ts": user["expiration_ts"]})
    except Exception as e:
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))    



@router.post("/auth/refresh-token", response_model=TokenResponse)
async def refresh_token(data: RefreshTokenRequest):
    refresh_token = data.refreshToken
    try:
        payload = jwt.decode(refresh_token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username is None or payload.get("type") != "refresh":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid refresh token",
                headers={"WWW-Authenticate": "Bearer"},
            )
        new_access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        new_access_token = create_access_token(
            data={"sub": username}, expires_delta=new_access_token_expires
        )
        user_sessions[username] = new_access_token

        return TokenResponse(code=0, success=True, data=TokenData(
            accessToken=new_access_token,
            refreshToken=refresh_token,
            expires=datetime.utcnow() + new_access_token_expires
        ))
    except ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Refresh token expired",
            headers={"WWW-Authenticate": "Bearer"},
        )
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid refresh token",
            headers={"WWW-Authenticate": "Bearer"},
        )
    except Exception as e:
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=str(e),
            headers={"WWW-Authenticate": "Bearer"},
        )


# 获取用户列表
@router.post("/auth/get_user_list")
async def get_user_list(msg: GetUserMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        await check_super_admin_role(current_user)
        
        search_params = msg.search_params
        page = msg.page
        limit = msg.limit
        
        username = None
        nickname = None
        phone = None
        
        # 解析搜索参数
        if search_params:
            try:
                search_dict = json.loads(search_params)
                if 'username' in search_dict:
                    username = search_dict.get('username')
                if 'nickname' in search_dict:
                    nickname = search_dict.get('nickname')
                if 'phone' in search_dict:
                    phone = search_dict.get('phone')
                # 可以处理更多的搜索参数
            except json.JSONDecodeError:
                return HTTPException(status_code=400, detail="search_params的参数非JSON格式字符串，解析失败")

        
        # 构建基本查询
        query = select(users_table)

        # 根据是否有搜索参数添加过滤条件
        if search_params:
            # Check if 'device_name' is in search_params and perform substring search
            if username:
                username_pattern = f"%{username}%"
                query = query.where(users_table.c.username.ilike(username_pattern))

            # Check if 'project_name' is in search_params and perform substring search
            if nickname:
                nickname_pattern = f"%{nickname}%"
                query = query.where(users_table.c.nickname.ilike(nickname_pattern))

            # Check if 'sn' is in search_params and perform substring search
            if phone:
                phone_pattern = f"%{phone}%"
                query = query.where(users_table.c.phone.ilike(phone_pattern))

            if username == '' and nickname == '' and phone == '':
                # 查询数据总数
                total_query = select(func.count()).select_from(users_table)
                total_count = await database.fetch_val(total_query)
            else:
                # 查询数据总数
                total_query = query.with_only_columns(func.count()).order_by(None)
                total_count = await database.fetch_val(total_query)
        else:
            total_query = select(func.count()).select_from(users_table)
            total_count = await database.fetch_val(total_query)

        # 根据数据量决定是返回全部还是分页
        if total_count <= limit * (page - 1):
            results = await database.fetch_all(query.order_by(users_table.c.id))
        else:
            offset = (page - 1) * limit
            results = await database.fetch_all(query.order_by(users_table.c.id).offset(offset).limit(limit))

        if results:
            users = [
                UserListData(
                    id=result["id"],
                    avatar=result["avatar"],
                    username=result["username"],
                    nickname=result["nickname"],
                    roles=[role_name_converter(role) for role in result["roles"].split(",")],
                    email=result["email"],
                    phone=result["phone"],
                    expiration_ts=format_datetime(result["expiration_ts"]),
                    create_ts=format_datetime(result["create_ts"]),
                    last_login_ts=format_datetime(result["last_login_ts"]),
                    disabled="是" if result["disabled"] else "否"
                ) for result in results
            ]

            logger.info(f"查询users_table表成功")
            return UserListResult(code=0, success=True, count=total_count, data=users, other={"total_count": total_count})
        else:
            logger.info(f"查询users_table表 无数据")
            return UserListResult(code=0, success=True, count=0, data=[], other={"total_count": 0})
    except Exception as e:
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))    

# 新增用户
@router.post("/auth/add_user")
async def add_user(msg: AddUserMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        await check_super_admin_role(current_user)
        
        log_msg = ""

        # 检查用户名是否已存在
        query = select(users_table).where(users_table.c.username == msg.username)
        user = await database.fetch_one(query)
        if user:
            raise HTTPException(status_code=409, detail=f"账号：{msg.username} 已存在")

        # 保存发送和接收的消息到数据库
        query = users_table.insert().values(
            avatar="https://avatar.vercel.sh/rauchg.svg?text=Admin",
            username=msg.username,
            nickname=msg.nickname,
            roles=msg.roles,
            hashed_password=hash_password(msg.password),
            email=msg.email,
            phone=msg.phone,
            expiration_ts=datetime.strptime(msg.expiration_ts, '%Y-%m-%d').date(),
            create_ts=datetime.utcnow(),
            last_login_ts=datetime.utcnow(),
            disabled=msg.disabled
        )
        rows_affected = await transaction_manager.execute(query)
        
        log_msg = f"添加用户:{msg.username} 成功"
        await insert_msg_to_logs_table(current_user, log_msg)
        
        return CommonResult(code=0, success=True, data={"msg": log_msg})
    except HTTPException as e:
        # 如果是HTTPException，则直接抛出
        raise e
    except Exception as e:
        logger.error(traceback.format_exc())
        logger.error(f"添加用户失败：{e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/auth/add_users")
async def add_users(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_active_user)
):
    try:
        await check_super_admin_role(current_user)

        # 读取和解析 XLSX 文件
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active
        
        # 检查第一行标题
        header = [cell.value for cell in sheet[1]]
        if header[0] != "账号" or header[1] != "密码" or header[2] != "昵称" or header[3] != "角色" or \
            header[4] != "邮箱" or header[5] != "手机号" or header[6] != "禁用" or header[7] != "到期时间":
            raise HTTPException(status_code=400, detail="Excel 表格格式不正确")
        
        users = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 第一行是标题行
            username, password, nickname, roles, email, phone, disabled, expiration_ts = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]  # 只提取前n列
            # 不能为空
            if not username or not password or not roles or not disabled or not expiration_ts:
                continue
            
            query = select(users_table).where(users_table.c.username == username)
            user_resp = await database.fetch_one(query)
            if user_resp:
                return CommonResult(code=0, success=False, data={"msg":f"上传失败，用户名：{username} 已存在"})
            # 是否还需要其他限制条件？
            # query = select(projects_table.c.batch_num).where(projects_table.c.project_name == project_name)
            # project = await database.fetch_one(query)
            # if not project:
            #     raise HTTPException(status_code=404, detail=f"上传失败，项目名称：{project_name} 不存在")

            users.append({
                "avatar": "https://avatar.vercel.sh/rauchg.svg?text=Admin",
                "username": username,
                "nickname": nickname,
                "roles": role_name_converter(roles, 1),
                "hashed_password": hash_password(password),
                "email": email,
                "phone": phone,
                "expiration_ts": datetime.strptime(expiration_ts, '%Y-%m-%d').date(),
                "create_ts": datetime.utcnow(),
                "last_login_ts": datetime.utcnow(),
                "disabled": 1 if disabled == "是" else 0,
            })
        
        query = users_table.insert().values(users)
        rows_affected = await transaction_manager.execute(query)
        
        log_msg = "批量新增用户数据 上传成功，共添加了 {} 条数据".format(rows_affected)
        await insert_msg_to_logs_table(current_user, log_msg)
        
        return CommonResult(code=0, success=True, data={"msg": log_msg})
    except Exception as e:
        logger.error(traceback.format_exc())
        logger.error(f"批量添加用户失败： {e}")
        raise HTTPException(status_code=500, detail=str(e))

# 删除用户
@router.post("/auth/del_user")
async def del_user(msg: DelUserMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        await check_super_admin_role(current_user)

        # 检查用户名是否已存在
        query = select(users_table).where(users_table.c.id == msg.id)
        user = await database.fetch_one(query)
        if not user:
            raise HTTPException(status_code=409, detail=f"账号ID：{msg.id} 不存在，请检查此账号是否已经被他人删除")

        # 删除存在的用户ID
        delete_query = delete(users_table).where(users_table.c.id == msg.id)
        rows_affected = await transaction_manager.execute(delete_query)
        
        log_msg = f"删除用户ID：{msg.id} 成功"
        await insert_msg_to_logs_table(current_user, log_msg)

        # 返回结果
        return CommonResult(code=0, success=True, data={"msg": log_msg})
    except SQLAlchemyError as e:
        logger.error(f"删除用户失败： {e}")
        raise HTTPException(status_code=500, detail="数据库操作失败")
    except Exception as e:
        logger.error(f"删除用户失败： {e}")
        raise HTTPException(status_code=500, detail=str(e))
 
# 批量删除用户
@router.post("/auth/del_users")
async def del_users(msgs: DelUsersMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        await check_super_admin_role(current_user)
        
        ids = [msg.id for msg in msgs.ids]
        if not ids:
            raise HTTPException(status_code=400, detail="用户ID列表不能为空")

        # 查询存在的用户ID
        query = select(users_table.c.id).where(users_table.c.id.in_(ids))
        existing_ids = await database.fetch_all(query)
        existing_ids = [proj['id'] for proj in existing_ids]

        # 找到不存在的用户ID
        non_existing_ids = set(ids) - set(existing_ids)

        # 删除存在的用户ID
        if existing_ids:
            delete_query = delete(users_table).where(users_table.c.id.in_(existing_ids))
            rows_affected = await transaction_manager.execute(delete_query)

        log_msg = f"删除了用户ID为{existing_ids}的用户"
        await insert_msg_to_logs_table(current_user, log_msg)

        # 返回结果
        return CommonResult(
            code=0,
            success=True,
            data={
                "msg": "批量删除完成",
                "deleted_ids": existing_ids,
                "non_existing_ids": list(non_existing_ids)
            }
        )
    except SQLAlchemyError as e:
        logger.error(f"删除用户失败： {e}")
        raise HTTPException(status_code=500, detail="数据库操作失败")
    except Exception as e:
        logger.error(f"删除用户失败： {e}")
        raise HTTPException(status_code=500, detail=str(e))
 
 
# 修改用户信息（超级管理员）
@router.post("/auth/update_user")
async def update_user(msg: UpdateUserMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        query = select(users_table).where(
            and_(
                users_table.c.id == msg.id,
                users_table.c.username == msg.ori_username,
            )
        )
        user_info = await database.fetch_one(query)
        if not user_info:
            raise HTTPException(status_code=400, detail=f"账号：{msg.ori_username} 不存在，请检查此账号是否已经被他人删除")

        if msg.username != msg.ori_username:
            # 账号不能重复
            query = select(users_table).where(users_table.c.username == msg.username)
            user_info = await database.fetch_one(query)
            if user_info:
                raise HTTPException(status_code=400, detail=f"账号：{msg.username} 已存在，请重新输入")

        query = (
            update(users_table)
            .where(users_table.c.id == msg.id)
            .values(
                username=msg.username,
                nickname=msg.nickname,
                roles=role_name_converter(msg.roles, 2),
                email=msg.email,
                phone=msg.phone,
                expiration_ts=datetime.strptime(msg.expiration_ts, '%Y-%m-%d').date(),
                disabled=1 if msg.disabled == "是" else 0,
            )
        )
        rows_affected = await transaction_manager.execute(query)
        
        log_msg = f'更新用户信息 原账号：{msg.ori_username} 新账号：{msg.username}，昵称：{msg.nickname}，角色：{msg.roles}，邮箱：{msg.email}，电话：{msg.phone}，到期时间：{msg.expiration_ts}，是否禁用：{msg.disabled}'
        await insert_msg_to_logs_table(current_user, log_msg)
        
        return CommonResult(code=0, success=True, data={"msg": log_msg})
    except Exception as e:
        logger.error(f"更新设备失败： {e}")
        raise HTTPException(status_code=500, detail=str(e))
 

# 修改密码
@router.post("/auth/update_password")
async def update_password(msg: UpdatePasswordMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        resp_json = await get_user_roles(current_user["username"])
        if resp_json['code'] != 200:
            raise HTTPException(status_code=403, detail="获取当前用户权限失败，请检查数据库是否正常！")
        
        is_super_admin = False
        roles = resp_json['data']['roles']
        for role in roles:
            # 如果当前用户是超级管理员，则可以添加用户
            if role == "super_admin":
                is_super_admin = True

        # 如果不是超级管理员
        if not is_super_admin:
            # 如果当前用户修改的不是自己的密码，则抛出异常
            if current_user["username"] != msg.username:
                raise HTTPException(status_code=400, detail="用户名不匹配，您不能修改其他用户的密码")
        
            # 判断旧密码是否正确
            query = select(users_table).where(users_table.c.username == msg.username)
            user = await database.fetch_one(query)
            if not user or not verify_password(msg.old_password, user["hashed_password"]):
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="错误的 账号或密码",
                    headers={"WWW-Authenticate": "Bearer"},
                )
        
        
        new_hashed_password = hash_password(msg.new_password)
        query = (
            update(users_table)
            .where(users_table.c.username == msg.username)
            .values(
                hashed_password=new_hashed_password,
            )
        )
        
        rows_affected = await transaction_manager.execute(query)
        try:
            if rows_affected == 0:
                log_msg = f"无记录需要更新，没有username为{msg.username}的用户"
                logger.warning(log_msg)
                await insert_msg_to_logs_table(current_user, log_msg)
                return CommonResult(code=0, success=False, data={"msg": log_msg})
            else:
                log_msg = f"用户{msg.username}更新密码成功"
                logger.info(log_msg)
                await insert_msg_to_logs_table(current_user, log_msg)
                return CommonResult(code=0, success=True, data={"msg": log_msg})
            
        except Exception as e:
            log_msg = f"用户{msg.username} 更新密码失败，数据库操作执行异常"
            logger.error(f"用户{msg.username} 更新密码失败，{e}")
            await insert_msg_to_logs_table(current_user, log_msg)
            return CommonResult(code=0, success=False, data={"msg": log_msg})
        
    except Exception as e:
        logger.error(traceback.format_exc())
        logger.error(f"更新密码失败： {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
# 修改用户信息
@router.post("/auth/update_user_info")
async def update_user_info(msg: UpdateUserInfoMessage, current_user: dict = Depends(get_current_active_user)):
    try:
        
        
        query = select(users_table).where(users_table.c.username == msg.username)
        user = await database.fetch_one(query)
        if not user:
            logger.error(f"用户不存在，users_table没有username为{msg.username}的用户")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Incorrect username",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        query = (
            update(users_table)
            .where(users_table.c.username == msg.username)
            .values(
                username=msg.new_username,
                phone=msg.phone,
                email=msg.email,
            )
        )
        
        rows_affected = await transaction_manager.execute(query)
        try:
            if rows_affected == 0:
                log_msg = f"无记录需要更新，没有username为{msg.username}的用户"
                logger.warning(log_msg)
                await insert_msg_to_logs_table(current_user, log_msg)
                return CommonResult(code=0, success=False, data={"msg": log_msg})
            else:
                log_msg = f'更新用户信息 原账号：{msg.username}, 新账号：{msg.new_username} 电话：{msg.phone} 邮箱：{msg.email}'
                logger.info(log_msg)
                await insert_msg_to_logs_table(current_user, log_msg)
                return CommonResult(code=0, success=True, data={"msg": log_msg})
            
        except Exception as e:
            logger.error(f"更新用户信息失败，数据库操作执行异常, {e}")
            return CommonResult(code=0, success=False, data={"msg": "更新用户信息失败，数据库操作执行异常"})
        
    except Exception as e:
        logger.error(traceback.format_exc())
        logger.error(f"更新用户信息失败： {e}")
        raise HTTPException(status_code=500, detail=str(e))
    